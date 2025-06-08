# Step 2: Generate Excel Report with openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def generate_excel_report(summary):
    wb = load_workbook("templates/report_template.xlsx")
    ws = wb["Summary"]

    # Clear old data
    for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Insert new data
    for i, row in enumerate(summary.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=row.Region)
        ws.cell(row=i, column=2, value=row.Amount)

    # Add chart
    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(summary))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(summary))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "Sales by Region"
    ws.add_chart(chart, "D2")

    excel_output = "output/report_2025-05.xlsx"
    wb.save(excel_output)
    print("Excel report created:", excel_output)
